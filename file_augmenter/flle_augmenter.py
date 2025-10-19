from openpyxl import load_workbook, Workbook
from itertools import cycle, islice

source_path = "loans.xlsx"
target_path = "expanded_loans.xlsx"
target_rows = 100_000

src_wb = load_workbook(source_path, data_only=True, read_only=True)
src_ws = src_wb.active

n_cols = src_ws.max_column


def fixed_rows(ws, n_cols):

    for row in ws.iter_rows(min_row=1, max_col=n_cols, values_only=True):

        if row is None:
            yield tuple("" for _ in range(n_cols))
        else:
            r = tuple("" if v is None else v for v in row)  # keep empty strings for Excel
            if len(r) < n_cols:
                r = r + tuple("" for _ in range(n_cols - len(r)))
            elif len(r) > n_cols:
                r = r[:n_cols]
            yield r


rows_iter = fixed_rows(src_ws, n_cols)
header = next(rows_iter)
data_rows = list(rows_iter)

dst_wb = Workbook(write_only=True)
dst_ws = dst_wb.create_sheet(title="Sheet1")

dst_ws.append(header)

counter = 0
for row in islice(cycle(data_rows), target_rows):
    counter = counter + 1
    print(counter)
    dst_ws.append(row)

dst_wb.save(target_path)
