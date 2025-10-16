from __future__ import annotations

import abc
import ast
from dataclasses import dataclass, field
from datetime import date
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Literal, Dict, Any

from openpyxl import load_workbook

BORROWER_MAP = {
    "borrower id": "borrower_id",
    "birth year": "birth_year",
    "gender": "gender",
    "marital status": "marital_status",
    "children": "children",
    "residential status": "residential_status",
    "education": "education",
    "occupation": "occupation",
    "months at current employer": "months_at_current_employer",
    "employment status": "employment_status",
    "years working total": "years_working_total",
    "borrower income": "borrower_income",
    "borrower liabilities": "borrower_liabilities",
    "spouse income": "spouse_income",
    "spouse liabilities": "spouse_liabilities",
    "family income": "family_income",
    "family liabilities": "family_liabilities",
    "dti": "dti",
}

LOAN_MAP = {
    "loan id": "loan_id",
    "credit score": "credit_score",
    "loan amount": "loan_amount",
    "disbursal date": "disbursal_date",
    "interest rate": "interest_rate",
    "loan term": "loan_term",
    "borrower type": "borrower_type",
    "loan type": "loan_type",
    "loan status": "loan_status",
    "purpose": "purpose",
}

REPAYMENT_MAP = {
    "monthly payment": "monthly_payment",
    "outstanding principal": "outstanding_principal",
    "repaid principal": "repaid_principal",
    "outstanding interest": "outstanding_interest",
    "repaid interest": "repaid_interest",
    "repayment date": "repayment_date",
    "expected repayment date": "expected_repayment_date",
    "last debt payment date": "last_debt_payment_date",
    "arrears": "arrears",
    "delay interest": "delay_interest",
    "days late": "days_late",
    "payments": "payments",
}

COMPANY_MAP = {
    "city": "city",
    "activity": "activity",
    "sector": "sector",
    "product": "product",
    "number of employees": "number_of_employees",
    "annual revenue": "annual_revenue",
    "annual profit": "annual_profit",
    "company type": "company_type",
    "company age (years)": "company_age_years",
    "company description": "company_description",
    "shareholders equity": "shareholders_equity",
}

COLLATERAL_MAP = {
    "appraisal date": "appraisal_date",
    "appraisal provider": "appraisal_provider",
    "collateral description": "collateral_description",
    "collateral market value": "collateral_market_value",
    "collateral name": "collateral_name",
    "collateral owner": "collateral_owner",
    "guarantor title": "guarantor_title",
}

Severity = Literal["ERROR", "WARN", "INFO", "CLEAN"]


@dataclass
class Issue:
    code: str
    severity: Severity
    field: Optional[str]
    message: str
    value: Optional[object] = None
    suggestion: Optional[str] = None


@dataclass
class LoanRecord:
    borrower: BorrowerInfo
    loan: LoanInfo
    repayment: RepaymentInfo
    company: Optional[CompanyInfo] = None
    collateral: Optional[CollateralInfo] = None
    issues: List[Issue] = field(default_factory=list, init=False)

    def validate(self) -> List[Dict[int, Issue]]:
        issues: List[Issue] = []
        issues += self.borrower.validate()
        issues += self.loan.validate()
        issues += self.repayment.validate()
        issues += self.company.validate()
        issues += self.collateral.validate()

        mapped_issues = {self.loan.loan_id: issues}
        return mapped_issues


@dataclass
class BorrowerInfo:
    borrower_id: int
    birth_year: Optional[int] = None
    gender: Optional[str] = None
    marital_status: Optional[str] = None
    children: Optional[int] = None
    residential_status: Optional[str] = None
    education: Optional[str] = None
    occupation: Optional[str] = None
    months_at_current_employer: Optional[int] = None
    employment_status: Optional[str] = None
    years_working_total: Optional[int] = None
    borrower_income: Optional[float] = None
    borrower_liabilities: Optional[float] = None
    spouse_income: Optional[float] = None
    spouse_liabilities: Optional[float] = None
    family_income: Optional[float] = None
    family_liabilities: Optional[float] = None
    dti: Optional[float] = None

    def validate(self) -> List[Issue]:
        issues: List[Issue] = []

        for field in ["borrower_income", "spouse_income", "family_income",
                      "borrower_liabilities", "spouse_liabilities", "family_liabilities",
                      "children", "months_at_current_employer", "years_working_total"]:
            val = getattr(self, field)
            if val is not None and val < 0:
                issues.append(Issue("NEGATIVE_VALUE", "ERROR", field,
                                    "Field should not be negative.", val))
        if self.dti is not None:
            if self.dti < 0:
                issues.append(Issue("DTI_NEGATIVE", "ERROR", "dti", "DTI cannot be negative.", self.dti))

        return issues


@dataclass
class LoanInfo:
    loan_id: int
    credit_score: Optional[str] = None
    loan_amount: Optional[float] = None
    disbursal_date: Optional[date] = None
    interest_rate: Optional[float] = None
    loan_term: Optional[int] = None
    borrower_type: Optional[str] = None
    loan_type: Optional[str] = None
    loan_status: Optional[str] = None
    purpose: Optional[str] = None

    def validate(self) -> List[Issue]:
        issues: List[Issue] = []
        if self.loan_amount is not None and self.loan_amount <= 0:
            issues.append(Issue("AMOUNT_NONPOSITIVE", "ERROR", "loan_amount",
                                "Loan amount must be > 0.", self.loan_amount))

        date_fields = [f for f in self.__annotations__ if f.endswith("date")]
        for field in date_fields:
            val = getattr(self, field)
            if val == 'Not Valid date':
                issues.append(Issue("INVALID_DATE", "ERROR", "date issue", "Date is not valid formatted", val))

        return issues


@dataclass
class RepaymentInfo:
    monthly_payment: Optional[float] = None
    outstanding_principal: Optional[float] = None
    repaid_principal: Optional[float] = None
    outstanding_interest: Optional[float] = None
    repaid_interest: Optional[float] = None
    repayment_date: Optional[date] = None
    expected_repayment_date: Optional[date] = None
    last_debt_payment_date: Optional[date] = None
    arrears: Optional[float] = None
    delay_interest: Optional[float] = None
    days_late: Optional[int] = None
    payments: Optional[Dict[str, Any]] = None

    def validate(self) -> List[Issue]:
        issues: List[Issue] = []
        for field in ["monthly_payment", "outstanding_principal", "repaid_principal",
                      "outstanding_interest", "repaid_interest", "arrears", "delay_interest"]:
            val = getattr(self, field)
            if val is not None and val < 0:
                issues.append(Issue("NEGATIVE_VALUE", "ERROR", field,
                                    "Field should not be negative.", val))

        if self.days_late is not None and self.days_late < 0:
            issues.append(Issue("NEGATIVE_DAYS_LATE", "ERROR", "days_late",
                                "Days late cannot be negative.", self.days_late))

        if self.payments == "Not valid list of dicts":
            issues.append(Issue("NON_COMPLETE_PAYMENTS", "ERROR", "payments", "Payments column is not consistent"))

        if self.payments != 'Not valid list of dicts':
            for payment in self.payments:
                payment_date = payment.get('Payment date')
                repayment_date = payment.get('Repayment date')

                if not payment_date or not repayment_date:
                    continue

                try:
                    date_payment = datetime.strptime(payment_date, "%d/%m/%Y")
                    date_repayment = datetime.strptime(repayment_date, "%d/%m/%Y")
                    payment_diff = (date_payment - date_repayment).days
                    if payment_diff > 90:
                        issues.append(Issue("DEFAULT", "ERROR", "payments", f"Payment expired {payment_diff} days", f"payment date: {date_payment} -- repayment date:{date_repayment}"))
                except Exception as e:
                    issues.append(Issue("ParseError", "WARN", "Payment date", str(e), payment))

        date_fields = [f for f in self.__annotations__ if f.endswith("date")]
        for field in date_fields:
            val = getattr(self, field)
            if val == 'Not Valid date':
                issues.append(Issue("INVALID_DATE", "ERROR", field, "Date is not valid formatted"))

        return issues


@dataclass
class CompanyInfo:
    city: Optional[str] = None
    activity: Optional[str] = None
    sector: Optional[str] = None
    product: Optional[str] = None
    number_of_employees: Optional[int] = None
    annual_revenue: Optional[float] = None
    annual_profit: Optional[float] = None
    company_type: Optional[str] = None
    company_age_years: Optional[int] = None
    company_description: Optional[str] = None
    shareholders_equity: Optional[float] = None

    def validate(self) -> List[Issue]:
        issues: List[Issue] = []
        if self.number_of_employees is not None and self.number_of_employees < 0:
            issues.append(Issue("NEGATIVE_EMPLOYEES", "ERROR", "number_of_employees",
                                "Employee count cannot be negative.", self.number_of_employees))
        if self.annual_revenue is not None and self.annual_revenue < 0:
            issues.append(Issue("NEGATIVE_REVENUE", "ERROR", "annual_revenue",
                                "Revenue cannot be negative.", self.annual_revenue))
        return issues


@dataclass
class CollateralInfo:
    appraisal_date: Optional[date] = None
    appraisal_provider: Optional[str] = None
    collateral_description: Optional[str] = None
    collateral_market_value: Optional[float] = None
    collateral_name: Optional[str] = None
    collateral_owner: Optional[str] = None
    guarantor_title: Optional[str] = None

    def validate(self) -> List[Issue]:
        issues: List[Issue] = []
        if self.collateral_market_value is not None and self.collateral_market_value <= 0:
            issues.append(Issue("COLLATERAL_NONPOSITIVE", "ERROR", "collateral_market_value",
                                "Collateral market value must be > 0.", self.collateral_market_value))

        date_fields = [f for f in self.__annotations__ if f.endswith("date")]
        for field in date_fields:
            val = getattr(self, field)
            if val == 'Not Valid date':
                issues.append(Issue("INVALID_DATE", "ERROR", field, "Date is not valid formatted"))

        return issues


class LoanParser(abc.ABC):
    @abc.abstractmethod
    def parse_for(self, file_names: Path) -> List[LoanRecord]:
        ...


class XLSXLoanParser(LoanParser):

    def __init__(self):
        ...

    def parse_for(self, file_path: Path) -> List[LoanRecord]:
        loan_file = load_workbook(file_path, data_only=True, read_only=True).active
        header_cells = next(loan_file.iter_rows(min_row=1, max_row=1))
        headers = [self.__norm(str(c.value)) if c.value is not None else "" for c in header_cells]
        enumerated_headers = {i: h for i, h in enumerate(headers) if h}

        records: list[LoanRecord] = []

        for row in loan_file.iter_rows(min_row=2, values_only=True):
            row_dict = {enumerated_headers[i]: row[i] for i in enumerated_headers.keys()}

            borrower_kwargs = self.__extract(BORROWER_MAP, row_dict)
            loan_kwargs = self.__extract(LOAN_MAP, row_dict)
            repay_kwargs = self.__extract(REPAYMENT_MAP, row_dict)
            company_kwargs = self.__extract(COMPANY_MAP, row_dict)
            coll_kwargs = self.__extract(COLLATERAL_MAP, row_dict)

            borrower_id = borrower_kwargs.get("borrower_id")
            loan_id = loan_kwargs.get("loan_id")
            if not borrower_id or not loan_id:
                continue

            record = LoanRecord(
                borrower=BorrowerInfo(**borrower_kwargs),
                loan=LoanInfo(**loan_kwargs),
                repayment=RepaymentInfo(**repay_kwargs),
                company=CompanyInfo(**company_kwargs),
                collateral=CollateralInfo(**coll_kwargs),
            )
            records.append(record)

        return records

    def __extract(self, fields_map: Dict[str, str], row: Dict[str, Any]) -> Dict[str, Any]:
        out: Dict[str, Any] = {}
        for col_label, field in fields_map.items():
            val = row.get(col_label)
            if field.endswith("_date"):
                out[field] = self.__to_date(val)
            elif field == 'payments':
                out[field] = self.__to_list_of_dict(val)

            elif field in {"loan_amount", "monthly_payment", "outstanding_principal", "repaid_principal",
                           "outstanding_interest", "repaid_interest", "arrears", "delay_interest",
                           "annual_revenue", "annual_profit", "shareholders_equity", "interest_rate",
                           "borrower_income", "borrower_liabilities", "spouse_income",
                           "spouse_liabilities", "family_income", "family_liabilities",
                           "collateral_market_value", "dti"}:
                out[field] = self.__to_float(val)
            elif field in {"credit_score", "loan_term", "days_late",
                           "children", "months_at_current_employer",
                           "years_working_total", "number_of_employees",
                           "company_age_years", "birth_year"}:
                out[field] = self.__to_int(val)
            else:
                out[field] = None if val in (None, "") else val
        return out

    @staticmethod
    def __norm(s: str) -> str:
        return s.strip().lower()

    @staticmethod
    def __to_float(x):
        if x in (None, ""):
            return None
        try:
            if isinstance(x, str):
                x = x.replace(" ", "").replace(",", ".")
            return float(x)
        except (ValueError, TypeError):
            return "Not Valid float"

    @staticmethod
    def __to_int(x):
        if x in (None, ""):
            return None
        try:
            if isinstance(x, float) and x.is_integer():
                return int(x)
            return int(str(x).strip())
        except (ValueError, TypeError):
            return "Not Valid int"

    @staticmethod
    def __to_date(x):
        if x in (None, ""):
            return None
        if isinstance(x, date):
            return x
        if isinstance(x, datetime):
            return x.date()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S.%f"):
            try:
                return datetime.strptime(str(x).strip(), fmt).date()
            except (ValueError, TypeError):
                continue
        return "Not Valid date"

    @staticmethod
    def __to_list_of_dict(val):
        try:
            return ast.literal_eval(val)
        except Exception:
            return "Not valid list of dicts"
